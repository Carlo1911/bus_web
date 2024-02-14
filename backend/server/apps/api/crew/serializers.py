import os
from abc import abstractmethod
from datetime import datetime
from decimal import Decimal

from django.contrib.auth.models import Permission
from django.db.models.query_utils import Q
from openpyxl import load_workbook
from rest_framework import serializers

from api.users.serializers import CustomUserSerializer, BasicCustomUserSerializer
from crew.models import (
    Commission,
    CommissionLog,
    Motive,
    Personal,
    PersonalPosition,
    Position,
)
from users.models import (
    CustomUser,
    Role,
)


class MotiveSerializer(serializers.ModelSerializer):
    class Meta:
        model = Motive
        fields = "__all__"


class PositionSerializer(serializers.ModelSerializer):
    class Meta:
        model = Position
        fields = "__all__"


class PersonalPositionSerializer(serializers.ModelSerializer):
    position = PositionSerializer(read_only=True)

    class Meta:
        model = PersonalPosition
        fields = "__all__"


class CommissionLogSerializer(serializers.ModelSerializer):
    class Meta:
        model = CommissionLog
        fields = "__all__"


class CommissionSerializer(serializers.ModelSerializer):
    logs = CommissionLogSerializer(many=True, read_only=True)

    def validate(self, attrs):
        # TODO: Validar que el el usuario tenga un cargo que admita comisión
        date_from = attrs.get("date_from")
        date_to = attrs.get("date_to")
        percentage = attrs.get("percentage")
        # TODO: Validar que haya fechas
        if date_from > date_to:
            raise serializers.ValidationError({"date_from": "La fecha de inicio debe ser menor a la fecha de fin"})
        if percentage < 0 or percentage > 100:
            raise serializers.ValidationError({"percentage": "El porcentaje debe estar entre 0 y 100"})
        return attrs

    def create(self, validated_data):
        request = self.context.get("request")
        commission = Commission.objects.create(**validated_data)
        CommissionLog.objects.create(
            commission=commission,
            date_from=commission.date_from,
            date_to=commission.date_to,
            percentage=commission.percentage,
            user=request.user if request else None,
        )
        return commission

    def update(self, instance, validated_data):
        request = self.context.get("request")
        CommissionLog.objects.create(
            commission=instance,
            date_from=instance.date_from,
            date_to=instance.date_to,
            percentage=instance.percentage,
            user=request.user if request else None,
        )
        instance.date_from = validated_data.get("date_from", instance.date_from)
        instance.date_to = validated_data.get("date_to", instance.date_to)
        instance.percentage = validated_data.get("percentage", instance.percentage)
        instance.personal = validated_data.get("personal", instance.personal)
        instance.route = validated_data.get("route", instance.route)
        instance.status = validated_data.get("status", instance.status)
        instance.save()
        return instance

    class Meta:
        model = Commission
        fields = "__all__"


class SimplePersonalSerializer(serializers.ModelSerializer):
    class Meta:
        model = Personal
        fields = "__all__"


class PersonalUserSerializer(serializers.ModelSerializer):

    class Meta:
        model = Personal
        fields = "__all__"


class PersonalSerializer(serializers.ModelSerializer):
    user = CustomUserSerializer()
    motives = MotiveSerializer(many=True)
    positions = PositionSerializer(many=True)
    comissions = CommissionSerializer(source="commissions", many=True)

    class Meta:
        model = Personal
        fields = "__all__"


class CreatePersonalSerializer(serializers.ModelSerializer):
    user = CustomUserSerializer(read_only=True)
    password = serializers.CharField(
        write_only=True,
        required=True,
        style={"input_type": "password", "placeholder": "Contraseña"},
    )
    is_superuser = serializers.BooleanField(
        required=False, write_only=True, style={"placeholder": "Estado de superusuario"}
    )
    username = serializers.CharField(write_only=True, style={"placeholder": "Nombre de usuario"})
    first_name = serializers.CharField(write_only=True, style={"placeholder": "Nombre"})
    last_name = serializers.CharField(write_only=True, style={"placeholder": "Apellido Paterno"})
    maternal_last_name = serializers.CharField(write_only=True, style={"placeholder": "Apellido Materno"})
    email = serializers.EmailField(write_only=True, style={"placeholder": "Dirección de correo electrónico"})
    is_active = serializers.BooleanField(write_only=True, style={"placeholder": "Activo"})
    is_staff = serializers.BooleanField(write_only=True, style={"placeholder": "Es staff"})
    document_number = serializers.CharField(write_only=True, style={"placeholder": "RUT"})
    cellphone = serializers.CharField(write_only=True, style={"placeholder": "Número de celular"})
    phone = serializers.CharField(
        write_only=True,
        allow_blank=True,
        allow_null=True,
        style={"placeholder": "Número de casa"},
    )
    image = serializers.FileField(
        required=False,
        write_only=True,
        allow_null=True,
        allow_empty_file=True,
        style={"placeholder": "Imagen"},
    )
    sex = serializers.ChoiceField(write_only=True, choices=CustomUser.Sex.choices, style={"placeholder": "Sexo"})
    address = serializers.CharField(
        write_only=True,
        allow_blank=True,
        allow_null=True,
        style={"placeholder": "Dirección"},
    )
    work_phone = serializers.CharField(
        required=False,
        write_only=True,
        allow_blank=True,
        allow_null=True,
        style={"placeholder": "Número de trabajo"},
    )
    work_address = serializers.CharField(
        required=False,
        write_only=True,
        allow_blank=True,
        allow_null=True,
        style={"placeholder": "Dirección de celular"},
    )
    groups = serializers.PrimaryKeyRelatedField(
        many=True,
        queryset=Role.objects.all(),
        write_only=True,
        allow_empty=True,
        allow_null=True,
        style={"placeholder": "Roles"},
    )
    user_permissions = serializers.PrimaryKeyRelatedField(
        many=True,
        write_only=True,
        queryset=Permission.objects.all(),
        allow_null=True,
        allow_empty=True,
        style={"placeholder": "Permisos"},
    )
    status = serializers.ChoiceField(choices=Personal.Status.choices, style={"placeholder": "Estado"})
    contract_type = serializers.ChoiceField(
        choices=Personal.ContractType.choices, style={"placeholder": "Tipo de contrato"}
    )
    contract_date = serializers.DateField(allow_null=True, style={"placeholder": "Fecha de contrato"})
    date_from = serializers.DateField(required=False, allow_null=True, style={"placeholder": "Fecha de inicio"})
    date_to = serializers.DateField(required=False, allow_null=True, style={"placeholder": "Fecha de finalización"})
    motives = serializers.PrimaryKeyRelatedField(
        many=True,
        write_only=True,
        queryset=Motive.objects.all(),
        allow_null=True,
        allow_empty=True,
        style={"placeholder": "Motivo"},
    )
    positions = serializers.PrimaryKeyRelatedField(
        many=True,
        write_only=True,
        queryset=Position.objects.all(),
        allow_null=True,
        allow_empty=True,
        style={"placeholder": "Cargo"},
    )
    manage_personal = serializers.CharField(default=serializers.CurrentUserDefault())

    class Meta:
        model = Personal
        fields = [
            "user",
            "password",
            "is_superuser",
            "username",
            "first_name",
            "last_name",
            "maternal_last_name",
            "email",
            "is_staff",
            "is_active",
            "document_number",
            "cellphone",
            "phone",
            "image",
            "sex",
            "address",
            "work_phone",
            "work_address",
            "groups",
            "user_permissions",
            "status",
            "contract_type",
            "contract_date",
            "motives",
            "positions",
            "date_from",
            "date_to",
            "manage_personal",
        ]

    def validate(self, attrs):
        username = attrs.get("username")
        email = attrs.get("email")
        document_number = attrs.get("document_number")

        users = CustomUser.objects.filter(username=username)
        if self.instance is not None:
            users = users.exclude(pk=self.instance.pk)
        if users.exists():
            raise serializers.ValidationError({"username": "Ya existe un usuario con el mismo username"})

        users = CustomUser.objects.filter(email=email)
        if self.instance is not None:
            users = users.exclude(pk=self.instance.pk)
        if users.exists():
            raise serializers.ValidationError({"email": "Ya existe un usuario con el mismo email"})

        users = CustomUser.objects.filter(document_number=document_number)
        if self.instance is not None:
            users = users.exclude(pk=self.instance.pk)
        if users.exists():
            raise serializers.ValidationError({"document_number": "Ya existe un usuario con el mismo RUT"})

        date_from = attrs.get("date_from")
        date_to = attrs.get("date_to")
        if date_from and date_to and date_from > date_to:
            raise serializers.ValidationError("La fecha de inicio debe ser menor a la fecha de finalización")
        return attrs

    def create(self, validated_data):
        # TODO: Check what happens if I want to change a motive or a position
        groups = validated_data.pop("groups", [])
        user_permissions = validated_data.pop("user_permissions", [])
        motives = validated_data.pop("motives", [])
        positions = validated_data.pop("positions", [])
        date_from = validated_data.pop("date_from")
        date_to = validated_data.pop("date_to")
        manage_personal = validated_data.pop("manage_personal")
        personal_data = {
            "status": validated_data.pop("status"),
            "contract_type": validated_data.pop("contract_type"),
            "contract_date": validated_data.pop("contract_date"),
        }
        user = CustomUser.objects.create_user(**validated_data)
        user.add_groups(groups)
        user.add_permissions(user_permissions)
        personal_data["user"] = user
        personal = Personal.objects.create(**personal_data)
        personal.add_motives(motives, date_from, date_to, manage_personal)
        personal.add_positions(positions)
        return personal

    def update(self, instance, validated_data):
        instance.user.is_superuser = validated_data.get("is_superuser", instance.user.is_superuser)
        instance.user.username = validated_data.get("username", instance.user.username)
        instance.user.first_name = validated_data.get("first_name", instance.user.first_name)
        instance.user.last_name = validated_data.get("last_name", instance.user.last_name)
        instance.user.maternal_last_name = validated_data.get("maternal_last_name", instance.user.maternal_last_name)
        instance.user.email = validated_data.get("email", instance.user.email)
        instance.user.is_staff = validated_data.get("is_staff", instance.user.is_staff)
        instance.user.is_active = validated_data.get("is_active", instance.user.is_active)
        instance.user.document_number = validated_data.get("document_number", instance.user.document_number)
        instance.user.cellphone = validated_data.get("cellphone", instance.user.cellphone)
        instance.user.phone = validated_data.get("phone", instance.user.phone)
        instance.user.image = validated_data.get("image", instance.user.image)
        instance.user.sex = validated_data.get("sex", instance.user.sex)
        instance.user.address = validated_data.get("address", instance.user.address)
        instance.user.work_phone = validated_data.get("work_phone", instance.user.work_phone)
        instance.user.work_address = validated_data.get("work_address", instance.user.work_address)
        instance.user.save()
        instance.user.add_groups(validated_data.get("groups", instance.user.groups.all().values_list("id", flat=True)))
        instance.user.add_permissions(
            validated_data.get(
                "user_permissions",
                instance.user.user_permissions.all().values_list("id", flat=True),
            )
        )
        instance.status = validated_data.get("status", instance.status)
        instance.contract_type = validated_data.get("contract_type", instance.contract_type)
        instance.contract_date = validated_data.get("contract_date", instance.contract_date)
        instance.save()

        motives = instance.motives.all()
        motives_id = validated_data.get("motives", [])
        date_to = validated_data.get("date_to")
        date_from = validated_data.get("date_from")
        manage_personal = validated_data.get("manage_personal")
        for motive_id in motives_id:
            if motive_id not in motives.values_list("id", flat=True):
                instance.add_motives([motive_id], date_from, date_to, manage_personal)

        position_ids = validated_data.get("positions", [])
        # Clear positions
        instance.positions.clear()
        # Assign new positions
        for position_id in position_ids:
            instance.add_positions([position_id])

        return instance


class AssignPermissionSerializer(serializers.Serializer):
    permissions_id = serializers.PrimaryKeyRelatedField(
        many=True,
        queryset=Permission.objects.all(),
        write_only=True,
        style={"placeholder": "Permisos"},
    )

    @abstractmethod
    def create(self, validated_data):
        pass

    @abstractmethod
    def update(self, instance, validated_data):
        pass


class CheckUserSerializer(serializers.Serializer):
    document_number = serializers.CharField(
        max_length=12,
        required=False,
        allow_null=True,
        allow_blank=True,
        style={"placeholder": "RUT"},
    )
    email = serializers.EmailField(
        required=False,
        allow_null=True,
        allow_blank=True,
        style={"placeholder": "Correo electrónico"},
    )
    exists = serializers.BooleanField(read_only=True)
    message = serializers.CharField(read_only=True)

    def validate(self, attrs):
        attrs["exists"] = False
        attrs["message"] = ""
        if not attrs.get("document_number") and not attrs.get("email"):
            attrs["exists"] = False
            attrs["message"] = "Debe ingresar un RUT o un correo electrónico"
        personals = Personal.objects.filter(
            Q(user__document_number=attrs.get("document_number")) | Q(user__email=attrs.get("email"))
        )
        if personals.exists():
            attrs["exists"] = True
            attrs["message"] = "El usuario ya existe"
        return attrs

    @abstractmethod
    def create(self, validated_data):
        pass

    @abstractmethod
    def update(self, instance, validated_data):
        pass


class MassiveCommissionPositionSerializer(serializers.Serializer):
    file = serializers.FileField(
        required=True,
        write_only=True,
        allow_null=False,
        allow_empty_file=False,
        style={"placeholder": "Archivo"},
    )

    def validate(self, attrs):
        xlsfile = attrs.get("file")
        _, extension = os.path.splitext(xlsfile.name)
        # TODO: Add csv support
        if extension not in [".xlsx", ".xls"]:
            raise serializers.ValidationError({"file": "Archivo cargado no tiene extension de excel"})
        # TODO: check if personal_id is valid
        # TODO: check if personal don't have an active commission
        return attrs

    def process_file(self):
        file = self.validated_data.get("file")
        workbook = load_workbook(filename=file, read_only=True)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)
        for row in worksheet.iter_rows(min_row=2, max_col=4, values_only=True):
            if row[0] is not None:
                commission_data = {
                    "personal_id": int(row[0]),
                    "route_id": int(row[1]),
                    "percentage": Decimal(row[2]),
                    "date_from": datetime.strptime(row[3], "%d/%m/%Y").date(),
                    "date_to": datetime.strptime(row[4], "%d/%m/%Y").date(),
                }

                commission, created = Commission.objects.get_or_create(
                    personal=commission_data["personal_id"],
                    route=commission_data["route_id"],
                    defaults=commission_data,
                )
                CommissionLog.objects.create(
                    commission=commission,
                    date_from=commission.date_from,
                    date_to=commission.date_to,
                    percentage=commission.percentage,
                )
                if not created:
                    commission.percentage = commission_data["percentage"]
                    commission.date_from = commission_data["date_from"]
                    commission.date_to = commission_data["date_to"]
                    commission.save()

    @abstractmethod
    def create(self, validated_data):
        pass

    @abstractmethod
    def update(self, instance, validated_data):
        pass


class BasicPersonalSerializer(serializers.ModelSerializer):
    user = BasicCustomUserSerializer()

    class Meta:
        model = Personal
        fields = ['pk', 'user', 'status']
