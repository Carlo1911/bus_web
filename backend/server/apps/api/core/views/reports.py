from datetime import (
    date,
    datetime,
)
from decimal import Decimal

from django.db.models.aggregates import Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import (
    OpenApiParameter,
    extend_schema,
)
from openpyxl import Workbook
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.parsers import JSONParser
from rest_framework.permissions import IsAuthenticatedOrReadOnly
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet

from api.core.serializers import (
    AdvanceSalesReportRequestSerializer,
    AdvanceSalesResponseSerializer,
    CommissionReportRequestSerializer,
    CommissionReportResponseSerializer,
    DashboardSaleIncomeRequestSerializer,
    PaxListSimpleRequestSerializer,
    SalesReportRequestSerializer,
    SalesReportResponseSerializer,
)
from core.models import (
    Sale,
    Service,
    Ticket,
)
from core.utils import gerenate_csv_from_dict
from crew.models import Commission
from weasyprint import (
    CSS,
    HTML,
)
from weasyprint.text.fonts import FontConfiguration


class ReportViewSet(GenericViewSet):
    queryset = Sale.objects.all().prefetch_related("tickets").order_by("id")
    permission_classes = (IsAuthenticatedOrReadOnly,)
    parser_classes = (JSONParser,)
    ordering_fields = ()
    filter_class = None

    def _sales_report_map(self, services, source: str):
        sales_report_data = []
        for service in services:
            if ticket := service.tickets.first():
                sale = ticket.sales.first()
                refunded_ticket_status = [Ticket.TicketStatus.REFUNDED, Ticket.TicketStatus.PARTIAL_REFUNDED]
                returned_amount = (
                    service.tickets.filter(ticket_status__in=refunded_ticket_status, ticket_source=source)
                    .aggregate(returned_amount=Sum("price"))
                    .get("returned_amount")
                )

                returned_ticket_count = service.tickets.filter(
                    ticket_status__in=refunded_ticket_status, ticket_source=source
                ).count()

                sales_report_data.append(
                    {
                        "service": service.name,
                        "start_city": service.start_city.name,
                        "end_city": service.end_city.name,
                        "bus": service.bus.name,
                        "total_seats": service.bus.layout.total_seats,
                        "date": service.datetime_publish.strftime("%d/%m/%Y %H:%M"),
                        "total_passenger": service.tickets.filter(ticket_source=source).count(),
                        "total": sale.cash_register.actual_amount if sale.cash_register is not None else Decimal(0),
                        "returned_amount": returned_amount,
                        "returned_ticket_count": returned_ticket_count,
                    }
                )
        return sales_report_data

    @extend_schema(
        summary="report sales",
        description="report sales",
        tags=["Reports"],
        request=DashboardSaleIncomeRequestSerializer,
        responses={("200", "text/csv"): OpenApiTypes.STR},
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="end_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="cash_registers",
                type={"type": "array", "items": {"type": "integer"}},
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="source",
                type=OpenApiTypes.STR,
                enum=["web", "branch_office"],
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(detail=False, methods=["get"], url_path="sales", url_name="sales")
    def sales(self, request):
        serializer = SalesReportRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            start_date: datetime = validated_data.get("start_date")
            end_date = validated_data.get(
                "end_date",
                datetime(start_date.year, start_date.month, start_date.day, 23, 59, 59),
            )
            cash_registers: [int] = validated_data.get("cash_registers")
            source: str = validated_data.get("source")

            services = self.filter_sales_report(cash_registers, end_date, source, start_date)
            sales_report_data = self._sales_report_map(services, source)
            csv_header = (
                "Servicio",
                "Ciudad Inicio",
                "Ciudad Fin",
                "Bus",
                "N° de asientos",
                "Fecha",
                "Total pasajeros",
                "Total",
                "Total devoluciones",
                "Tickets devueltos",
            )
            return gerenate_csv_from_dict(csv_header, "Sales_Report.csv", sales_report_data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="report sales detail",
        description="report sales detail",
        tags=["Reports"],
        request=SalesReportRequestSerializer,
        responses=SalesReportResponseSerializer(many=True),
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="end_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="cash_registers",
                type={"type": "array", "items": {"type": "integer"}},
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="source",
                type=OpenApiTypes.STR,
                enum=["web", "branch_office"],
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(detail=False, methods=["get"], url_path="sales-detail", url_name="sales-detail")
    def sales_detail(self, request):
        serializer = SalesReportRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            start_date: datetime = validated_data.get("start_date")
            end_date = validated_data.get(
                "end_date",
                datetime(start_date.year, start_date.month, start_date.day, 23, 59, 59),
            )
            cash_registers: [int] = validated_data.get("cash_registers")
            source: str = validated_data.get("source")

            services = self.filter_sales_report(cash_registers, end_date, source, start_date)
            page = self.paginate_queryset(services)
            if page is not None:
                data = self._sales_report_map(page, source)
                serializer = SalesReportResponseSerializer(data, many=True)
                return self.get_paginated_response(serializer.data)

            sales_report_data = self._sales_report_map(services, source)
            sales_report_serializer = SalesReportResponseSerializer(sales_report_data, many=True)
            return Response(sales_report_serializer.data, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def filter_sales_report(self, cash_registers, end_date, source, start_date):
        sales = Sale.objects.exclude(status__in=[Sale.Status.PENDING, Sale.Status.CANCELED]).filter(
            payment_date__gte=start_date,
            payment_date__lt=end_date,
            tickets__ticket_source=source,
        )
        if cash_registers:
            sales = sales.filter(cash_register_id__in=cash_registers)
        services_id = sales.values_list("tickets__service_id", flat=True)
        return Service.objects.filter(id__in=services_id)

    def _advanced_sales_map(self, sales):
        sales_report_data = []
        for sale in sales:
            if ticket := sale.tickets.first():
                sales_report_data.append(
                    {
                        "service": ticket.service.name,
                        "start_city": ticket.service.start_city.name,
                        "end_city": ticket.service.end_city.name,
                        "bus": ticket.service.bus.name,
                        "total_seats": ticket.service.bus.layout.total_seats,
                        "date": ticket.service.datetime_publish.strftime("%d/%m/%Y %H:%M"),
                        "total_passenger": sale.tickets.count(),
                        "total": sale.cash_register.actual_amount,
                    }
                )
        return sales_report_data

    @extend_schema(
        summary="report advance sales",
        description="report advance sales",
        tags=["Reports"],
        request=AdvanceSalesReportRequestSerializer,
        responses={("200", "text/csv"): OpenApiTypes.STR},
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="end_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="cash_registers",
                type={"type": "array", "items": {"type": "integer"}},
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="source",
                type=OpenApiTypes.STR,
                enum=["web", "branch_office"],
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="advance-sales",
        url_name="advance-sales",
    )
    def advance_sales(self, request):
        serializer = AdvanceSalesReportRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            start_date: datetime = validated_data.get("start_date")  # + timedelta(days=1)
            end_date: datetime = validated_data.get(
                "end_date", datetime(start_date.year, start_date.month, start_date.day, 23, 59, 59)
            )
            cash_registers: [int] = validated_data.get("cash_registers")
            source: str = validated_data.get("source")

            sales = Sale.objects.filter(
                tickets__service__datetime_publish__gte=start_date,
                tickets__service__datetime_publish__lt=end_date,
                tickets__ticket_source=source,
            )

            if cash_registers:
                sales = sales.filter(cash_register_id__in=cash_registers)

            services_id = sales.values_list("tickets__service_id", flat=True)
            services = Service.objects.filter(id__in=services_id)
            sales_report_data = self._sales_report_map(services, source)

            csv_header = (
                "Servicio",
                "Ciudad Inicio",
                "Ciudad Fin",
                "Bus",
                "N° de asientos",
                "Fecha",
                "Total pasajeros",
                "Total",
                "Total devoluciones",
                "Tickets devueltos",
            )
            return gerenate_csv_from_dict(csv_header, "Advance_Sales_Report.csv", sales_report_data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="report advance sales detail",
        description="report advance sales detail",
        tags=["Reports"],
        request=AdvanceSalesReportRequestSerializer,
        responses=AdvanceSalesResponseSerializer(many=True),
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="end_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="cash_registers",
                type={"type": "array", "items": {"type": "integer"}},
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="source",
                type=OpenApiTypes.STR,
                enum=["web", "branch_office"],
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="advance-sales-detail",
        url_name="advance-sales-detail",
    )
    def advance_sales_detail(self, request):
        serializer = AdvanceSalesReportRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            start_date: datetime = validated_data.get("start_date")  # + timedelta(days=1)
            end_date: datetime = validated_data.get(
                "end_date", datetime(start_date.year, start_date.month, start_date.day, 23, 59, 59)
            )

            cash_registers: [int] = validated_data.get("cash_registers")
            source: str = validated_data.get("source")

            sales = Sale.objects.filter(
                tickets__service__datetime_publish__gte=start_date,
                tickets__service__datetime_publish__lt=end_date,
                tickets__ticket_source=source,
            )

            if cash_registers:
                sales = sales.filter(cash_register_id__in=cash_registers)

            services_id = sales.values_list("tickets__service_id", flat=True)
            services = Service.objects.filter(id__in=services_id)

            page = self.paginate_queryset(services)
            if page is not None:
                data = self._sales_report_map(services, source)
                serializer = AdvanceSalesResponseSerializer(data, many=True)
                return self.get_paginated_response(serializer.data)

            sales_report_data = self._sales_report_map(services, source)
            sales_report_serializer = AdvanceSalesResponseSerializer(sales_report_data, many=True)
            return Response(sales_report_serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def _commission_map(self, comissions):
        commission_data = []
        for commission in comissions:
            sales_amount = Decimal(0)
            sales_amount_ext = (
                Sale.objects.filter(seller=commission.personal)
                .aggregate(returned_amount=Sum("tickets__price"))
                .get("returned_amount")
            )
            if sales_amount_ext:
                sales_amount = sales_amount_ext

            commission_amount = commission.percentage * sales_amount / 100
            commission_data.append(
                {
                    "date_from": commission.date_from.strftime("%d/%m/%Y"),
                    "date_to": commission.date_to.strftime("%d/%m/%Y"),
                    "route": commission.route.name,
                    "branch_office": commission.personal.branch.name,
                    "document_number": commission.personal.user.document_number,
                    "personal": commission.personal.user.get_full_name(),
                    "percentage": commission.percentage,
                    "commission_amount": commission_amount,
                    "sales_amount": sales_amount,
                    "status": "Activo" if commission.status == Commission.Status.ACTIVE else "Inactivo",
                }
            )
        return commission_data

    @extend_schema(
        summary="report commission",
        description="report commission",
        tags=["Reports"],
        request=CommissionReportRequestSerializer,
        responses={("200", "text/csv"): OpenApiTypes.STR},
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="end_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(detail=False, methods=["get"], url_path="commission", url_name="commission")
    def commission(self, request):
        serializer = CommissionReportRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            start_date: date = validated_data.get("start_date")  # + timedelta(days=1)
            end_date: date = validated_data.get("end_date", date(start_date.year, start_date.month, start_date.day))

            commissions = Commission.objects.filter(date_from__gte=start_date, date_to__lte=end_date)
            commission_data = self._commission_map(commissions)

            csv_header = (
                "Desde",
                "Hasta",
                "Ruta",
                "Sucursal",
                "RUT",
                "Nombre",
                "Porcentaje",
                "Monto comisión",
                "Estado",
            )
            return gerenate_csv_from_dict(csv_header, "commission_report.csv", commission_data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="report commission detail",
        description="report commission detail",
        tags=["Reports"],
        request=CommissionReportRequestSerializer,
        responses=CommissionReportResponseSerializer(many=True),
        parameters=[
            OpenApiParameter(
                name="start_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
            OpenApiParameter(
                name="end_date",
                type=OpenApiTypes.DATETIME,
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(detail=False, methods=["get"], url_path="commission-detail", url_name="commission-detail")
    def commission_detail(self, request):
        serializer = CommissionReportRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            start_date: date = validated_data.get("start_date")  # + timedelta(days=1)
            end_date: date = validated_data.get("end_date", date(start_date.year, start_date.month, start_date.day))

            commissions = Commission.objects.filter(date_from__gte=start_date, date_to__lte=end_date)
            page = self.paginate_queryset(commissions)
            if page is not None:
                data = self._commission_map(page)
                serializer = CommissionReportResponseSerializer(data, many=True)
                return self.get_paginated_response(serializer.data)

            commission_data = self._commission_map(commissions)
            return Response(
                CommissionReportResponseSerializer(commission_data, many=True).data, status=status.HTTP_200_OK
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="Pax List Simple",
        description="Pax List Simple PDF",
        tags=["Reports"],
        request=PaxListSimpleRequestSerializer,
        responses={("200", "application/pdf"): OpenApiTypes.BINARY},
        parameters=[
            OpenApiParameter(
                name="service",
                type=OpenApiTypes.INT32,
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(detail=False, methods=["get"], url_path="pax-list-simple", url_name="pax-list-simple")
    def pax_list_simple(self, request):
        serializer = PaxListSimpleRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            service_id = validated_data.get("service")
            service = Service.objects.get(pk=service_id)
            tickets = service.tickets.all()
            drivers = []
            assistants = []
            passengers = []
            total_amount = Decimal(0)

            for ticket in tickets:
                passengers.append(
                    {
                        "seat": ticket.seat.code,
                        "start_city": service.start_city,
                        "end_city": service.end_city,
                        "price": ticket.price,
                    }
                )
                total_amount += ticket.price

            for driver in service.drivers.all():
                drivers.append(
                    {
                        "name": driver.user.get_full_name(),
                    }
                )

            for assistant in service.assistants.all():
                assistants.append(
                    {
                        "name": assistant.user.get_full_name(),
                    }
                )

            all_seats = []
            for seat in service.web_seats.exclude(kind="space"):
                all_seats.append(seat.code)

            seats = []
            for seat in service.sales_seats.exclude(kind="space"):
                for ticket in tickets:
                    if seat != ticket.seat and seat.kind != "":
                        seats.append(seat.code)

            data = {
                "start_city": service.start_city,
                "end_city": service.end_city,
                "service_name": service.name,
                "travel_date": service.datetime_publish.strftime("%d/%m/%Y"),
                "end_travel_date": service.datetime_publish.strftime("%d/%m/%Y"),
                # TODO: Obtener duración por tramos
                "travel_time": service.datetime_publish.strftime("%H:%M"),
                "end_travel_time": service.datetime_publish.strftime("%H:%M"),  # TODO: Obtener duración por tramos
                "start_bus_stop": service.route.start_bus_stop,
                "end_bus_stop": service.route.end_bus_stop,
                "arrival_time": service.last_arrival_datetime.strftime("%H:%M"),
                "available_seats": ", ".join(seats) if len(seats) > 0 else ", ".join(all_seats),
                "drivers": drivers,
                "assistants": assistants,
                "passengers": passengers,
                "bus": service.bus,
                "total_amount": total_amount,
                "available_seats_count": len(seats) if len(seats) > 0 else len(all_seats),
                "tickets_count": len(tickets),
                "total_passengers": len(tickets),
                "printer_user": request.user.get_username(),
                "print_dater": datetime.now().strftime("%d/%m/%Y %H:%M"),
            }

            page_length = 297
            css_string = f"""
                        @page {{
                            margin: 0;
                            size: 70mm {page_length}mm;
                        }}
                        body {{font-size: 11px;}}
                        body ul {{ margin: 0;}}
                        .space {{margin-top: 10px;}}
                        .passangers-title
                        .destiny-title
                        .origin-title {{
                            font-size: 13px;
                        }}
                        """
            font_config = FontConfiguration()
            css = CSS(
                string=css_string,
                font_config=font_config,
            )

            html = render_to_string("pax-list/pax-list-simple.html", data)
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f"inline; pax_list_simple_{service.id}.pdf"
            HTML(string=html).write_pdf(response, stylesheets=[css], font_config=font_config)
            return response

    @extend_schema(
        summary="Pax List Transport ministery",
        description="Pax List Transport Ministery PDF",
        tags=["Reports"],
        request=PaxListSimpleRequestSerializer,
        responses={("200", "application/pdf"): OpenApiTypes.BINARY},
        parameters=[
            OpenApiParameter(
                name="service",
                type=OpenApiTypes.INT32,
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(detail=False, methods=["get"], url_path="pax-list-transport", url_name="pax-list-transport")
    def pax_list_transport_ministry(self, request):
        serializer = PaxListSimpleRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            service_id = validated_data.get("service")
            header_excel = [
                "PPU Bus",
                "Nombre Empresa",
                "Región Inscripción",
                "Folio Inscripción",
                "Ciudad Origen",
                "Fecha Salida",
                "Hora Salida",
                "Ciudad Destino",
                "Fecha Llegada",
                "Hora Llegada",
                "N° Asiento",
                "Tipo Documento (RUN, Pasaporte, Identificación Local)",
                "NÚMERO DE DOCUMENTO DE VIAJE",
                "Nombre Pasajero",
                "Pais de Origen",
                "Celular",
                "Correo Electrónico",
            ]

            wb = Workbook()
            ws = wb.active

            for i, header in enumerate(header_excel):
                ws.cell(row=1, column=i + 1).value = header

            service = Service.objects.get(pk=service_id)
            tickets = service.tickets.all()

            passengers = [
                {
                    "PPU": service.bus.plate,
                    "Nombre Empresa": "Andimar",
                    "Región Inscripción": "O¨higgins",
                    "Folio Inscripción": "800199",
                    "Ciudad Origen": ticket.start_bus_stop.name,
                    "Fecha Salida": service.datetime_publish.strftime("%d/%m/%Y"),
                    "Hora Salida": service.datetime_publish.strftime("%H:%M"),
                    "Ciudad Destino": ticket.end_bus_stop.name,
                    "Fecha Llegada": service.last_arrival_datetime.strftime("%d/%m/%Y"),
                    # TODO: Obtener hora de llegada calculada del tramo
                    "Hora Llegada": service.last_arrival_datetime.strftime("%H:%M"),
                    "N° Asiento": ticket.seat.code,
                    "Tipo Documento": ticket.passenger.user.document_type,
                    "NÚMERO DE DOCUMENTO DE VIAJE": ticket.passenger.user.document_number,
                    "Nombre": ticket.passenger.user.get_full_name(),
                    "Pais de Origen": ticket.passenger.user.nationality.name,
                    "Celular": ticket.passenger.user.cellphone,
                    "Correo Electrónico": ticket.passenger.user.email,
                }
                for ticket in tickets
            ]

            for i, passenger in enumerate(passengers):
                for j, value in enumerate(passenger.values()):
                    ws.cell(row=i + 2, column=j + 1).value = value

            formatted_name = (
                f"pax_list_ministerio_{service.route.start_bus_stop.name}_{service.route.end_bus_stop.name}.xlsx"
            )
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; \
                                filename="{formatted_name}'
                },
            )
            wb.save(response)

            return response
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="Pax List Full",
        description="Pax List Full PDF",
        tags=["Reports"],
        request=PaxListSimpleRequestSerializer,
        responses={("200", "application/pdf"): OpenApiTypes.BINARY},
        parameters=[
            OpenApiParameter(
                name="service",
                type=OpenApiTypes.INT32,
                location=OpenApiParameter.QUERY,
            ),
        ],
    )
    @action(detail=False, methods=["get"], url_path="pax-list-full", url_name="pax-list-full")
    def pax_list_full(self, request):
        serializer = PaxListSimpleRequestSerializer(data=request.query_params)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            service_id = validated_data.get("service")
            service = Service.objects.get(pk=service_id)
            tickets = service.tickets.all()
            services_messages = service.messages.all()
            passengers = []
            pax_messages = []
            total_amount = Decimal(0)

            for message in services_messages:
                print_pax_message = {
                    'name': f'{message.personal.user.first_name} {message.personal.user.last_name}',
                    'document_number': message.personal.user.document_number,
                    'message': message.message
                }
                pax_messages.append(print_pax_message)

            for ticket in tickets:
                passengers.append(
                    {
                        "seat": ticket.seat.code,
                        "name": ticket.passenger.user.get_full_name(),
                        "document_number": ticket.passenger.user.document_number.replace(".", ""),
                        "cellphone": ticket.passenger.user.cellphone.replace("+56", ""),
                    }
                )
                total_amount += ticket.price

            drivers = [
                {
                    "name": driver.user.get_full_name(),
                }
                for driver in service.drivers.all()
            ]

            assistants = [
                {
                    "name": assistant.user.get_full_name(),
                }
                for assistant in service.assistants.all()
            ]

            all_seats = [seat.code for seat in service.web_seats.exclude(kind="space")]

            seats = []
            for seat in service.sales_seats.exclude(kind="space"):
                seats.extend(seat.code for ticket in tickets if seat != ticket.seat and seat.kind != "")

            data = {
                "start_city": service.start_city,
                "end_city": service.end_city,
                "service_name": service.name,
                "travel_date": service.datetime_publish.strftime("%d/%m/%Y"),
                "end_travel_date": service.datetime_publish.strftime("%d/%m/%Y"),
                "travel_time": service.datetime_publish.strftime("%H:%M"),
                # TODO: Obtener duración por tramos
                "pax_messages": pax_messages,
                "end_travel_time": service.datetime_publish.strftime("%H:%M"),
                "start_bus_stop": service.route.start_bus_stop,
                "end_bus_stop": service.route.end_bus_stop,
                "arrival_time": service.last_arrival_datetime.strftime("%H:%M"),
                "available_seats": ", ".join(seats) if len(seats) > 0 else ", ".join(all_seats),
                "drivers": drivers,
                "assistants": assistants,
                "passengers": passengers,
                "bus": service.bus,
                "total_amount": total_amount,
                "tickets": tickets,
                "available_seats_count": len(seats),
                "tickets_count": len(tickets),
                "total_passengers": len(tickets),
                "printer_user": request.user.get_username(),
                "print_dater": datetime.now().strftime("%d/%m/%Y %H:%M"),
            }

            page_length = 297
            css_string = f"""
                            @page {{
                                margin: 0;
                                size: 70mm {page_length}mm;
                            }}
                            body {{font-size: 11px;}}
                            body ul {{ margin: 0;}}
                            .space {{margin-top: 10px;}}
                            .passangers-title
                            .destiny-title
                            .origin-title {{
                                font-size: 13px;
                            }}
                            """
            font_config = FontConfiguration()
            css = CSS(
                string=css_string,
                font_config=font_config,
            )

            html = render_to_string("pax-list/pax-list-full.html", data)
            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f"inline; pax_list_full_{service.id}.pdf"
            HTML(string=html).write_pdf(response, stylesheets=[css], font_config=font_config)
            return response
